'''
16/08/203
Pham Minh Nhat
PNM3HC
Get DID info from excel sheet

'''



import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import os
import re


D_ID = 'DID'
filename = 'DID_sheet'
filename_proposal = 'Create_function_VBA.xlsm'

work_sheet = 'Info_sheet'
work_sheet_proposal = 'proposal'



cwd = os.getcwd()

def find_ws(keyname,wbl):
    sheet_name = ''
    for i in wbl:
        print(i)
        if(keyname in i):
            sheet_name = i
    return sheet_name


def Find_file_Name(dir,keyname):
    fileNames = os.listdir(dir)
    file_name = ''
    for fileName in fileNames:
        if(keyname in fileName):
            file_name = fileName
    return file_name

# find_cell : find the cell contain keyword
def Find_cell(keyname,row_max,col_max_num,wsl):
    col_return = 0
    row_return = 0
    row_s = row_max
    col_s = col_max_num
    print(row_s,col_s)
    for row in range(1,row_max+1):
        for col in range(1,col_max_num+1):
            col_char = get_column_letter(col)
            if keyname in str(wsl[col_char +str(row)].value):
                col_return = col
                row_return = row
    print('[' + str(get_column_letter(col_return))+str(row_return) + '] : '+wsl[get_column_letter(col_return)+str(row_return)].value)
    return col_return,row_return

# get_cell : get the zone locate all cells (max column, max row in sheet)
def get_cell(extract_cell):
    for i in range (len(extract_cell)):
        if extract_cell[i] == '.':
            cell_c = extract_cell[i+1:-1]
            break
    num = re.findall(r'\d+', cell_c)
    char = cell_c.replace(num[0],'')
    row_max = int(num[0])
    col_max = int(column_index_from_string(char))
    return row_max, col_max

#======================================================================
#Access Excel proposal
start_time = time.time()


print('Accessing '+str(filename_proposal))
wbl_proposal = load_workbook(filename_proposal)
ws_name_proposal = find_ws(work_sheet_proposal,wbl_proposal.sheetnames)
if ws_name_proposal != '':
    wsl_proposal = wbl_proposal[ws_name_proposal]
else:
    wsl_proposal = wbl_proposal.active
print('accessing: '+ str(wsl_proposal))


# Find cell zone

extract_cell = str(wsl_proposal[1][-1])
row_max, col_max = get_cell(extract_cell)
extract_cell = str(wsl_proposal[get_column_letter(col_max)][-1])
row_max, col_max = get_cell(extract_cell)
print('row max in proposal : '+str(row_max)+'\tcolumn max in proposal : '+str(col_max))

DID_info_header = ['DID','Byte no','Phys','Func','RAMCELL','Phys','Hex','Phys','Hex','Phys',
                   'Hex','Phys','Hex','Phys','Hex','Phys','Hex','Phys','Hex','Phys','Hex','Cond check',
                   'Condition ','True Value Cond','False Value Cond','Condition2','Value Cond2','False Value Cond','Fixed value']
DID_info_proposal_total = []
DID_info_proposal_total.append(DID_info_header)
Ramcell_list = '[RAMCELL]'

#thru rows
above_DID = ''
for row in range(4,row_max):
    DID_info_proposal = []

    # DID
    if wsl_proposal['F'][row].value: 
        DID_name = wsl_proposal['F'][row].value
        above_DID = DID_name
    else:
        DID_name = above_DID
    DID_info_proposal.append(DID_name[-4:])

    # Byte no
    byte_pos = str(wsl_proposal['H'][row].value)
    start_pos = byte_pos.find('-')
    if start_pos :
        DID_info_proposal.append(byte_pos[0:start_pos])
    else:
        DID_info_proposal.append(byte_pos[0:])

    # Addr sp Phys
    DID_info_proposal.append('x')

    # Addr sp Func
    DID_info_proposal.append('')

    # Ramcell
    DID_info_proposal.append(wsl_proposal['G'][row].value)
    Ramcell_list = Ramcell_list + '\n' + str(wsl_proposal['G'][row].value)
    

    # go thru columns to get test data from column AA to AO
    start_col = int(column_index_from_string('AA'))
    test_data = []
    for i in range(0,int(column_index_from_string('AO')-column_index_from_string('AA')),2):
        cur_col = i + start_col
        if wsl_proposal[get_column_letter(cur_col)][row].value and wsl_proposal[get_column_letter(cur_col+1)][row].value:
            test_data.append(str(wsl_proposal[get_column_letter(cur_col)][row].value))
            test_data.append(str(wsl_proposal[get_column_letter(cur_col+1)][row].value))
    fil_count = 16 - len(test_data)
    if fil_count > 0 :
        for i in range(fil_count):
            test_data.append('')
    for i in range(16):
        DID_info_proposal.append(test_data[i])
    
    # Cond check
    DID_info_proposal.append('FALSE')

    # Condition
    DID_info_proposal.append('')

    # True Value
    DID_info_proposal.append('')

    # False Value
    DID_info_proposal.append('')

    # Condition2
    DID_info_proposal.append('')

    # Value Cond2
    DID_info_proposal.append('')

    # False Value Cond
    DID_info_proposal.append('')

    # Fixed value
    DID_info_proposal.append('')

    # Final step of row
    DID_info_proposal_total.append(DID_info_proposal)


# Create a new workbook
workbook = Workbook()

# Select the active sheet
sheet = workbook.active

# Iterate over the rows and columns of the data array
for row_index, row_data in enumerate(DID_info_proposal_total, start=1):
    for col_index, value in enumerate(row_data, start=1):
        # Write the value to the corresponding cell in the sheet
        sheet.cell(row=row_index, column=col_index, value=value)

# Save the workbook to a new Excel file
new_file = "DID_sheet.xlsx"
workbook.save(new_file)

# Specify the path and filename for the text file
file_path = "ramcell.lab"

# Open the file in write mode and save the string
with open(file_path, "w") as file:
    file.write(Ramcell_list)

#=======================================================================
#Access Excel Workbook

# File_dir = Find_file_Name(cwd,filename)

# print('Accessing '+str(File_dir))
# wbl = load_workbook(File_dir)
# ws_name = find_ws(work_sheet,wbl.sheetnames)
# if ws_name != '':
#     wsl = wbl[ws_name]
# else:
#     wsl = wbl.active
# print('accessing: '+ str(wsl))

# # Find cell zone

# extract_cell = str(wsl[1][-1])
# row_max, col_max = get_cell(extract_cell)
# extract_cell = str(wsl[get_column_letter(col_max)][-1])
# row_max, col_max = get_cell(extract_cell)
# print('row max : '+str(row_max)+'\tcolumn max : '+str(col_max))

# # get info tab

# for value in wsl.iter_rows(1,1,1,col_max,values_only=True):
#     header = value
#     print("header "+ str(header))
# Condition_pos = []
# RAMCELL_pos = []

# for i in range(len(header)):
#     if 'Condition' in header[i]:
#         Condition_pos.append(i+1)
#     if 'RAMCELL' in header[i]:
#         RAMCELL_pos.append(i+1)
#     if 'Cond check' in header[i]:
#         Cond_check_pos = i+1
#     if 'Fixed value' in header[i]:
#         Cond_val_pos = i+1

# # thru row

# DID_info_total = []
# for row in range(1,row_max):
#     DID_info = []
#     DID_info.append(wsl['A'][row].value)
#     DID_info.append(wsl['B'][row].value)
#     Addr_sp = [0,0]
#     if wsl['C'][row].value != None:
#         Addr_sp[0] = 1
#     if wsl['D'][row].value != None:
#         Addr_sp[1] = 1
#     DID_info.append(Addr_sp)

#     # DID handle
#     print('Row :'+str(row))
#     RAMCELL_info_row = []
#     RAMCELL_info_phys = []
#     RAMCELL_info_hex = []
#     RAMCELL_info = []
#     RAMCELL_info.append(wsl[get_column_letter(RAMCELL_pos[0])][row].value)
#     RAMCELL_pos_current = RAMCELL_pos[0]
#     for col in range(RAMCELL_pos[0]+1,Cond_check_pos):
#         if col not in RAMCELL_pos[1:]:
#             if wsl[get_column_letter(col)][row].value == None:
#                 break
#             else:
#                 if ((col-RAMCELL_pos_current)%2 == 1):
#                     if float(wsl[get_column_letter(col)][row].value) not in RAMCELL_info_phys:
#                         RAMCELL_info_phys.append(float(wsl[get_column_letter(col)][row].value))
#                 else:
#                     if '{0}{1}'.format('0'*(len(str(wsl[get_column_letter(col)][row].value))%2),str(wsl[get_column_letter(col)][row].value)) not in RAMCELL_info_hex:
#                         RAMCELL_info_hex.append('{0}{1}'.format('0'*(len(str(wsl[get_column_letter(col)][row].value))%2),str(wsl[get_column_letter(col)][row].value)))
#         else:
#             if wsl[get_column_letter(col)][row].value == None:
#                 break
#             else:
#                 RAMCELL_info.append(RAMCELL_info_phys)
#                 RAMCELL_info_phys = []
#                 RAMCELL_info.append(RAMCELL_info_hex)
#                 RAMCELL_info_hex = []
#                 RAMCELL_info_row.append(RAMCELL_info)
#                 RAMCELL_info = []
#                 RAMCELL_info.append(wsl[get_column_letter(col)][row].value)
#                 RAMCELL_pos_current = col
#     RAMCELL_info.append(RAMCELL_info_phys)
#     RAMCELL_info.append(RAMCELL_info_hex)
#     RAMCELL_info_row.append(RAMCELL_info)
#     DID_info.append(RAMCELL_info_row)
    
#     # Condition handle
#     if 'True' in str(wsl[get_column_letter(Cond_check_pos)][row].value):
#         DID_info.append(True)
#         print('condition check')
#         Condition_info_row = []
#         Condition_info = []
#         Condition_value = []
#         Condition_info.append(wsl[get_column_letter(Condition_pos[0])][row].value)
#         print('Condition row 1')
#         print(Condition_info_row)
#         for col in range(Condition_pos[0]+1,Cond_val_pos):
#             if col not in Condition_pos[1:]:
#                 if wsl[get_column_letter(col)][row].value != None:
#                     Condition_value.append(int(wsl[get_column_letter(col)][row].value))
#             else:
#                 print(Condition_pos)
#                 if wsl[get_column_letter(col)][row].value == None:
#                     break
#                 else:
#                     Condition_info.append(Condition_value)
#                     Condition_info_row.append(Condition_info)
#                     print('Condition row 2')
#                     print(Condition_info_row)
#                     Condition_info = []
#                     Condition_value = []
#                     Condition_info.append(wsl[get_column_letter(col)][row].value)
#         Condition_info.append(Condition_value)
#         Condition_info_row.append(Condition_info)
#         print('Condition row 3')
#         print(Condition_info_row)
#         Condition_value_hex_out = []
#         Condition_value_hex = str(wsl[get_column_letter(Cond_val_pos)][row].value)
#         Condition_value_hex = '{0}{1}'.format('0'*(len(Condition_value_hex)%2),Condition_value_hex)
#         Condition_value_hex_out.append(Condition_value_hex)
#         Condition_value_hex_out.append('XX')
#         Condition_info_row.append(Condition_value_hex_out)
#         print('Condition row 4')
#         print(Condition_info_row)
#         DID_info.append(Condition_info_row)
#     else:
#         print('No condition')
#         DID_info.append(False)
#     # test
#     print(DID_info)
#     DID_info_total.append(DID_info)
# print(DID_info_total)
# print('\n')


#================================================================================



print('\n\n\n\n===============================Completed==================================')
end_time = time.time()
print('Execution time:'+str(end_time-start_time)+' s')
