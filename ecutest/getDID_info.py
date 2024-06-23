"""
Created on Fri Aug 17 10:00:2023

@author: PNM3HC

"""



import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import os
import re


D_ID = 'DID'
filename = 'DID_sheet'
work_sheet = 'Info_sheet'



# cwd = os.getcwd()

def find_ws(keyname,wbl):
    sheet_name = ''
    for i in wbl:
        print(i)
        if(keyname in i):
            sheet_name = i
    return sheet_name


def Find_file_Name(dir,keyname):
    fileNames = os.listdir(dir)
    for fileName in fileNames:
        if(keyname in fileName):
            file_name = fileName
    return file_name

def Find_cell(keyname,row_max,col_max_num,wsl):
    col_return = 0
    row_return = 0
    row_s = row_max
    col_s = col_max_num
    print(row_s,col_s)
    for row in range(1,row_max+1):
        for col in range(1,col_max_num+1):
            col_char = get_column_letter(col)
            if keyname in str(wsl[col_char +str(row)].value):
                col_return = col
                row_return = row
    print('[' + str(get_column_letter(col_return))+str(row_return) + '] : '+wsl[get_column_letter(col_return)+str(row_return)].value)
    return col_return,row_return

def get_cell(extract_cell):
    for i in range (len(extract_cell)):
        if extract_cell[i] == '.':
            cell_c = extract_cell[i+1:-1]
            break
    num = re.findall(r'\d+', cell_c)
    char = cell_c.replace(num[0],'')
    row_max = int(num[0])
    col_max = int(column_index_from_string(char))
    return row_max, col_max
#Access Excel Workbook

def Get_info(File_dir):

    start_time = time.time()
    print('Accessing '+str(File_dir))
    wbl = load_workbook(File_dir)
    ws_name = find_ws(work_sheet,wbl.sheetnames)
    if ws_name != '':
        wsl = wbl[ws_name]
    else:
        wsl = wbl.active
    print('accessing: '+ str(wsl))

    # Find cell zone

    extract_cell = str(wsl[1][-1])
    row_max, col_max = get_cell(extract_cell)
    extract_cell = str(wsl[get_column_letter(col_max)][-1])
    row_max, col_max = get_cell(extract_cell)
    print('row max : '+str(row_max)+'\tcolumn max : '+str(col_max))

    # get info tab

    for value in wsl.iter_rows(1,1,1,col_max,values_only=True):
        header = value
    print(header)
    Condition_pos = []
    RAMCELL_pos = []

    for i in range(len(header)):
        if 'Condition' in header[i]:
            Condition_pos.append(i+1)
        if 'RAMCELL' in header[i]:
            RAMCELL_pos.append(i+1)
        if 'Cond check' in header[i]:
            Cond_check_pos = i+1
        if 'Fixed value' in header[i]:
            Cond_val_pos = i+1

    # thru row

    DID_info_total = []
    for row in range(1,row_max):
        DID_info = []
        DID_info.append(wsl['A'][row].value)
        DID_info.append(wsl['B'][row].value)
        Addr_sp = [0,0]
        if wsl['C'][row].value != None:
            Addr_sp[0] = 1
        if wsl['D'][row].value != None:
            Addr_sp[1] = 1
        DID_info.append(Addr_sp)

        # DID handle
        print('Row :'+str(row))
        RAMCELL_info_row = []
        RAMCELL_info_phys = []
        RAMCELL_info_hex = []
        RAMCELL_info = []
        RAMCELL_info.append(wsl[get_column_letter(RAMCELL_pos[0])][row].value)
        RAMCELL_pos_current = RAMCELL_pos[0]
        check_not_dup = True
        for col in range(RAMCELL_pos[0]+1,Cond_check_pos):
            if col not in RAMCELL_pos[1:]:
                if wsl[get_column_letter(col)][row].value == None:
                    break
                else:
                    
                    if ((col-RAMCELL_pos_current)%2 == 1):
                        if float(wsl[get_column_letter(col)][row].value) not in RAMCELL_info_phys:
                            RAMCELL_info_phys.append(float(wsl[get_column_letter(col)][row].value))
                            check_not_dup = True
                        else:
                            check_not_dup = False
                    else:
                        if check_not_dup:
                            RAMCELL_info_hex.append('{0}{1}'.format('0'*(len(str(wsl[get_column_letter(col)][row].value))%2),str(wsl[get_column_letter(col)][row].value)))
            else:
                if wsl[get_column_letter(col)][row].value == None:
                    break
                else:
                    RAMCELL_info.append(RAMCELL_info_phys)
                    RAMCELL_info_phys = []
                    RAMCELL_info.append(RAMCELL_info_hex)
                    RAMCELL_info_hex = []
                    RAMCELL_info_row.append(RAMCELL_info)
                    RAMCELL_info = []
                    RAMCELL_info.append(wsl[get_column_letter(col)][row].value)
                    RAMCELL_pos_current = col
        RAMCELL_info.append(RAMCELL_info_phys)
        RAMCELL_info.append(RAMCELL_info_hex)
        RAMCELL_info_row.append(RAMCELL_info)
        DID_info.append(RAMCELL_info_row)
        
        # Condition handle
        if 'TRUE' in str(wsl[get_column_letter(Cond_check_pos)][row].value).upper():
            DID_info.append(True)
            Condition_info_row = []
            Condition_info = []
            Condition_value = []
            Condition_info.append(wsl[get_column_letter(Condition_pos[0])][row].value)
            for col in range(Condition_pos[0]+1,Cond_val_pos):
                if col not in Condition_pos[1:]:
                    if wsl[get_column_letter(col)][row].value != None:
                        Condition_value.append(int(wsl[get_column_letter(col)][row].value))
                else:
                    if wsl[get_column_letter(col)][row].value == None:
                        break
                    else:
                        Condition_info.append(Condition_value)
                        Condition_info_row.append(Condition_info)
                        Condition_info = []
                        Condition_value = []
                        Condition_info.append(wsl[get_column_letter(col)][row].value)
            Condition_info.append(Condition_value)
            Condition_info_row.append(Condition_info)
            Condition_value_hex_out = []
            Condition_value_hex = str(wsl[get_column_letter(Cond_val_pos)][row].value)
            Condition_value_hex = '{0}{1}'.format('0'*(len(Condition_value_hex)%2),Condition_value_hex)
            Condition_value_hex_out.append(Condition_value_hex)
            Condition_value_hex_out.append('XX')
            Condition_info_row.append(Condition_value_hex_out)
            DID_info.append(Condition_info_row)
        else:
            DID_info.append(False)
        # test
        print(DID_info)
        DID_info_total.append(DID_info)
    # print(DID_info_total)
    print('\n')
    end_time = time.time()
    print(str(end_time-start_time)+' s')
    return DID_info_total