import os
import datetime
import openpyxl
import re

def detect_misra_violations(file_path):
    violations = []

    # Rule 2.1
    for line_number, line in enumerate(open(file_path, 'r'), start=1):
        if re.search(r'\bfor\s*\(.*;.*;.*\)', line):
            violations.append({
                'Rule': '2.1',
                'Line': line_number,
                'Detail': 'The essentials of the plain \'for\' statement shall not be bypassed.',
                'Solution': 'Review and modify the \'for\' loop to ensure it adheres to the rule.',
            })

    # Rule 2.2
    if '/*UNREACHABLE*/' in open(file_path).read():
        violations.append({
            'Rule': '2.2',
            'Line': None,
            'Detail': 'A project shall not contain unreachable code.',
            'Solution': 'Identify and remove unreachable code segments in the project.',
        })

    # Rule 2.3
    if 'typedef' in open(file_path).read() and 'UNUSED_TYPE' in open(file_path).read():
        violations.append({
            'Rule': '2.3',
            'Line': None,
            'Detail': 'A project shall not contain unused type declarations.',
            'Solution': 'Remove unused type declarations from the project.',
        })

    # Rule 3.3
    with open(file_path, 'r') as file:
        source_code = file.read()

    ast_module = None
    try:
        import astroid
        ast_module = astroid
    except ImportError:
        import ast

    if ast_module:
        tree = ast_module.parse(source_code)
        for node in ast_module.walk(tree):
            if isinstance(node, (ast_module.For, ast_module.While, ast_module.If)):
                if not isinstance(node.body[0], (ast_module.IfExp, ast_module.With, ast_module.Try, ast_module.If)):
                    if not isinstance(node.body[0], ast_module.stmt):
                        line_number = getattr(node, 'lineno', 0)
                        violations.append({
                            'Rule': '3.3',
                            'Line': line_number,
                            'Detail': 'The body of an iteration or selection statement should be a compound statement.',
                            'Solution': 'Enclose the body of the statement in curly braces {} to form a compound statement.',
                        })

    # Rule 3.4
    for line_number, line in enumerate(open(file_path, 'r'), start=1):
        if re.search(r'\bdo\s*{.*}\s*while\s*\(.*\)\s*;', line):
            violations.append({
                'Rule': '3.4',
                'Line': line_number,
                'Detail': 'The do-while statement shall be followed by a compound statement.',
                'Solution': 'Enclose the code block after the do-while statement in curly braces to form a compound statement.',
            })

    # Rule 3.5
    for line_number, line in enumerate(open(file_path, 'r'), start=1):
        if re.search(r'\bfor\s*\(.*;.+;.+\)\s*;', line):
            violations.append({
                'Rule': '3.5',
                'Line': line_number,
                'Detail': 'The three expressions of a for statement shall be present.',
                'Solution': 'Ensure that all three expressions of the for statement are present.',
            })

    # Rule 3.6
    for line_number, line in enumerate(open(file_path, 'r'), start=1):
        if re.search(r'\bwhile\s*\(.*\)\s*;', line):
            violations.append({
                'Rule': '3.6',
                'Line': line_number,
                'Detail': 'The while statement shall be followed by a compound statement.',
                'Solution': 'Enclose the code block after the while statement in curly braces to form a compound statement.',
            })

    # Rule 4.1
    for line_number, line in enumerate(open(file_path, 'r'), start=1):
        if re.search(r'\bif\s*\(.*\)\s*{.*}\s*else\s*{.*}\s*', line):
            violations.append({
                'Rule': '4.1',
                'Line': line_number,
                'Detail': 'The if-else if construct shall be terminated with an else statement.',
                'Solution': 'Add an else statement at the end of the if-else if construct.',
            })

    # Rule 4.2
    for line_number, line in enumerate(open(file_path, 'r'), start=1):
        if re.search(r'\bif\s*\(.*\)\s*{.*}\s*else\s*if\s*\(.*\)\s*{.*}\s*else\s*if\s*\(.*\)\s*{.*}\s*', line):
            violations.append({
                'Rule': '4.2',
                'Line': line_number,
                'Detail': 'There should be no more than one else if per branch.',
                'Solution': 'Remove the extra else if statements or restructure the code logic.',
            })

    # Rule 4.3
    for line_number, line in enumerate(open(file_path, 'r'), start=1):
        if re.search(r'\bif\s*\(.*\)\s*{.*}\s*else\s*{.*}\s*else\s*{.*}\s*', line):
            violations.append({
                'Rule': '4.3',
                'Line': line_number,
                'Detail': 'The else statement shall be followed by an if statement.',
                'Solution': 'Add an if statement after the else statement.',
            })

    # Rule 5.1
    for line_number, line in enumerate(open(file_path, 'r'), start=1):
        if re.search(r'\bgoto\s+\w+;', line):
            violations.append({
                'Rule': '5.1',
                'Line': line_number,
                'Detail': 'The goto statement shall not be used.',
                'Solution': 'Refactor the code to remove the use of goto statements.',
            })

    # Rule 5.2
    for line_number, line in enumerate(open(file_path, 'r'), start=1):
        if re.search(r'\bcontinue;', line):
            violations.append({
                'Rule': '5.2',
                'Line': line_number,
                'Detail': 'The continue statement shall not be used.',
                'Solution': 'Refactor the code to remove the use of continue statements.',
            })

    # Rule 5.3
    for line_number, line in enumerate(open(file_path, 'r'), start=1):
        if re.search(r'\bbreak;', line):
            violations.append({
                'Rule': '5.3',
                'Line': line_number,
                'Detail': 'The break statement shall not be used.',
                'Solution': 'Refactor the code to remove the useof break statements.',
            })

    # Rule 6.3
    with open(file_path, 'r') as file:
        source_code = file.read()

    bitwise_operators_regex = r'\b(<<|>>|&|\||\^)\b'
    bitwise_operators = re.findall(bitwise_operators_regex, source_code)
    for operator in bitwise_operators:
        line_number = source_code.count('\n', 0, source_code.index(operator)) + 1
        violations.append({
            'Rule': '6.3',
            'Line': line_number,
            'Detail': 'Bitwise operators should only be used on unsigned integer types.',
            'Solution': 'Ensure that bitwise operations are performed on unsigned integer types.',
        })

    # Rule 8.4
    for line_number, line in enumerate(open(file_path, 'r'), start=1):
        if re.search(r'\bwhile\s*\(.*\)\s*;', line):
            violations.append({
                'Rule': '8.4',
                'Line': line_number,
                'Detail': 'The while statement shall be followed by a compound statement.',
                'Solution': 'Enclose the code block after the while statement in curly braces to form a compound statement.',
            })

    # Rule 14.4
    for line_number, line in enumerate(open(file_path, 'r'), start=1):
        if re.search(r'\b#define\b.*\\$', line):
            violations.append({
                'Rule': '14.4',
                'Line': line_number,
                'Detail': 'The #define directive shall not be used to hide a macro expansion.',
                'Solution': 'Avoid using the backslash character to split macro definitions across multiple lines.',
            })
    # Rule 16.7
    for line_number, line in enumerate(open(file_path, 'r'), start=1):
        if re.search(r'\bsizeof\s*\(.*\)', line) and not re.search(r'\bsizeof\s*\(.*\s*\*\s*.*\)', line):
            violations.append({
                'Rule': '16.7',
                'Line': line_number,
                'Detail': 'The object passed to a sizeof operator shall not have an effective type that includes any variably modified type.',
                'Solution': 'Ensure that the sizeof operator is not applied to an object with a variably modified type.',
            })

    # Rule 18.4
    for line_number, line in enumerate(open(file_path, 'r'), start=1):
        if re.search(r'\bunion\b\s*{.*};', line):
            violations.append({
                'Rule': '18.4',
                'Line': line_number,
                'Detail': 'A union type declaration shall not contain members with overlapping sequences of bits.',
                'Solution': 'Ensure that the union declaration does not contain members with overlapping sequences of bits.',
            })

    # ... Add more rules as needed ...

    return violations

def export_to_excel(file_path, misra_violations):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set the column headers
    sheet["A1"] = "Line"
    sheet["B1"] = "MISRA ID"
    sheet["C1"] = "Detail"
    sheet["D1"] = "Solution"

    # Populate the data
    for row, violation in enumerate(misra_violations, start=2):
        sheet[f"A{row}"] = violation["Line"]
        sheet[f"B{row}"] = violation["MISRA ID"]
        sheet[f"C{row}"] = violation["Detail"]
        sheet[f"D{row}"] = violation["Solution"]

    # Save the Excel file
    workbook.save(file_path)

def process_c_files():
    current_folder = os.getcwd()
    files = [f for f in os.listdir(current_folder) if os.path.isfile(f) and f.endswith(".c")]

    misra_check_file = f"misra_check_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"
    workbook = openpyxl.Workbook()

    for file in files:
        misra_violations = detect_misra_violations(file)
        sheet = workbook.create_sheet(title=file)
        export_to_excel(sheet, misra_violations)

    workbook.save(misra_check_file)

process_c_files()