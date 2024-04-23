'''
16/08/203
Pham Minh Nhat
PNM3HC
Get DID info from excel sheet

'''



import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import os
import re


D_ID = 'DID'
filename = 'DID_sheet'
filename_proposal = 'Create_function_VBA'

work_sheet = 'Info_sheet'
work_sheet_proposal = 'proposal'



cwd = os.getcwd()

def find_ws(keyname,wbl):
    sheet_name = ''
    for i in wbl:
        print(i)
        if(keyname in i):
            sheet_name = i
    return sheet_name


def Find_file_Name(dir,keyname):
    fileNames = os.listdir(dir)
    for fileName in fileNames:
        if(keyname in fileName):
            file_name = fileName
    return file_name

# find_cell : find the cell contain keyword
def Find_cell(keyname,row_max,col_max_num,wsl):
    col_return = 0
    row_return = 0
    row_s = row_max
    col_s = col_max_num
    print(row_s,col_s)
    for row in range(1,row_max+1):
        for col in range(1,col_max_num+1):
            col_char = get_column_letter(col)
            if keyname in str(wsl[col_char +str(row)].value):
                col_return = col
                row_return = row
    print('[' + str(get_column_letter(col_return))+str(row_return) + '] : '+wsl[get_column_letter(col_return)+str(row_return)].value)
    return col_return,row_return

# get_cell : get the zone locate all cells (max column, max row in sheet)
def get_cell(extract_cell):
    for i in range (len(extract_cell)):
        if extract_cell[i] == '.':
            cell_c = extract_cell[i+1:-1]
            break
    num = re.findall(r'\d+', cell_c)
    char = cell_c.replace(num[0],'')
    row_max = int(num[0])
    col_max = int(column_index_from_string(char))
    return row_max, col_max

#======================================================================


def remove_characters_from_end(string, character, new_string):
    index = string.rfind(character)  # Find the last occurrence of the specified character
    if index != -1:  # If the character is found
        return (string[:index] + str(new_string))  # Return the string up to the character
    else:
        return string  # Return the original string if the character is not found

#Access Excel proposal
def Initial_info(File_dir_proposal):
    
    start_time = time.time()
    # File_dir_proposal = Find_file_Name(cwd,filename_proposal)

    print('Accessing '+str(File_dir_proposal))
    wbl_proposal = load_workbook(File_dir_proposal)
    ws_name_proposal = find_ws(work_sheet_proposal,wbl_proposal.sheetnames)
    if ws_name_proposal != '':
        wsl_proposal = wbl_proposal[ws_name_proposal]
    else:
        wsl_proposal = wbl_proposal.active
    print('accessing: '+ str(wsl_proposal))


    # Find cell zone

    extract_cell = str(wsl_proposal[1][-1])
    row_max, col_max = get_cell(extract_cell)
    extract_cell = str(wsl_proposal[get_column_letter(col_max)][-1])
    row_max, col_max = get_cell(extract_cell)
    print('row max in proposal : '+str(row_max)+'\tcolumn max in proposal : '+str(col_max))

    DID_info_header = ['DID','Byte no','Phys','Func','RAMCELL','Phys','Hex','Phys','Hex','Phys',
                    'Hex','Phys','Hex','Phys','Hex','Phys','Hex','Phys','Hex','Phys','Hex','Cond check',
                    'Condition ','True Value Cond','False Value Cond','Condition2','Value Cond2','False Value Cond','Fixed value']
    DID_info_proposal_total = []
    DID_info_proposal_total.append(DID_info_header)
    Ramcell_list = '[RAMCELL]'

    #thru rows
    above_DID = ''
    for row in range(4,row_max):
        DID_info_proposal = []

        # DID
        if wsl_proposal['F'][row].value: 
            DID_name = wsl_proposal['F'][row].value
            above_DID = DID_name
        else:
            DID_name = above_DID
        DID_info_proposal.append(DID_name[-4:])

        # Byte no
        byte_pos = str(wsl_proposal['H'][row].value)
        DID_info_proposal.append(byte_pos[0:1])

        # Addr sp Phys
        DID_info_proposal.append('x')

        # Addr sp Func
        DID_info_proposal.append('')

        # Ramcell
        DID_info_proposal.append(wsl_proposal['G'][row].value)
        Ramcell_list = Ramcell_list + '\n' + str(wsl_proposal['G'][row].value)
        

        # go thru columns to get test data from column AA to AO
        start_col = int(column_index_from_string('AA'))
        test_data = []
        for i in range(0,int(column_index_from_string('AO')-column_index_from_string('AA')),2):
            cur_col = i + start_col
            if wsl_proposal[get_column_letter(cur_col)][row].value and wsl_proposal[get_column_letter(cur_col+1)][row].value:
                test_data.append(str(wsl_proposal[get_column_letter(cur_col)][row].value))
                test_data.append(str(wsl_proposal[get_column_letter(cur_col+1)][row].value))
        fil_count = 16 - len(test_data)
        if fil_count > 0 :
            for i in range(fil_count):
                test_data.append('')
        for i in range(16):
            DID_info_proposal.append(test_data[i])
        
        # Cond check
        DID_info_proposal.append('FALSE')

        # Condition
        DID_info_proposal.append('')

        # True Value
        DID_info_proposal.append('')

        # False Value
        DID_info_proposal.append('')

        # Condition2
        DID_info_proposal.append('')

        # Value Cond2
        DID_info_proposal.append('')

        # False Value Cond
        DID_info_proposal.append('')

        # Fixed value
        DID_info_proposal.append('')

        # Final step of row
        DID_info_proposal_total.append(DID_info_proposal)


    # Create a new workbook
    workbook = Workbook()

    # Select the active sheet
    sheet = workbook.active

    # Iterate over the rows and columns of the data array
    for row_index, row_data in enumerate(DID_info_proposal_total, start=1):
        for col_index, value in enumerate(row_data, start=1):
            # Write the value to the corresponding cell in the sheet
            sheet.cell(row=row_index, column=col_index, value=value)

    # Save the workbook to a new Excel file
    new_file = "DID_sheet.xlsx"
    DID_file_name = remove_characters_from_end(File_dir_proposal,'\\',new_file)
    workbook.save(DID_file_name)

    # Specify the path and filename for the text file
    file_path = "ramcell.lab"

    # Open the file in write mode and save the string
    with open(file_path, "w") as file:
        file.write(Ramcell_list)

    #================================================================================

    end_time = time.time()
    print('Execution time:'+str(end_time-start_time)+' s')
    return 1
