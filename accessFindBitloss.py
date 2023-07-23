'''
09/08/2022
Pham Minh Nhat
PNM3HC
Find Factor & Norm, chech bitloss
'''



import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os
import numpy as np


'''
    You should use wb[sheetname]

    from openpyxl import load_workbook
    wb2 = load_workbook('test.xlsx')
    ws4 = wb2["New Title"]
    PS: You should check if your sheet in sheet names wb.sheetnames

    print(wb2.sheetnames)
    ['Sheet2', 'New Title', 'Sheet1']
'''
start_time = time.time()
cwd = os.getcwd()

def find_ws(keyname):
    sheet_name = ''
    for i in wbl.sheetnames :
        print(i)
        if(keyname in i):
            sheet_name = i
    return sheet_name


def Find_file_Name(dir,keyname):
    fileNames = os.listdir(dir)
    for fileName in fileNames:
        if(keyname in fileName):
            # file_path = os.path.abspath(os.path.join(dir, fileName))
            file_name = fileName

    return file_name

def Find_col(keyname,row_max,col_max_num):
    col_return = 0
    row_return = 0
    row_s = row_max
    col_s = col_max_num
    print(row_s,col_s)
    for row in range(1,row_max+1):
        for col in range(1,col_max_num+1):
            col_char = get_column_letter(col)
            # print(col_char)
            # print(str(wsl[col_char +str(row)].value))
            # print(wsl[col_char +str(row)].value)
            if keyname in str(wsl[col_char +str(row)].value):
                col_return = col
                row_return = row
                # print(col_return,row_return)
    print('[' + str(get_column_letter(col_return))+str(row_return) + '] : '+wsl[get_column_letter(col_return)+str(row_return)].value)
    return col_return,row_return

#Access Excel Workbook

Ramcell_dir = Find_file_Name(cwd,'Ramcell')
print('Accessing '+str(Ramcell_dir))
wbl = load_workbook(Ramcell_dir)
find_ws('analyze')
wsl = wbl.active
print('accessing: '+ str(wsl))



for row_count in range(1,1000):                                 # Count max rows and max cols 
    for j in range(1,1000):
        col_count = get_column_letter(j)
        if not wsl[col_count + str(row_count)].value == None:
            row_max = row_count
            col_max = col_count
            col_max_num = j
print('active rows: '+str(row_max))
print('active cols: '+str(col_max)+' : '+str(col_max_num))

DID_col,DID_row_1st = Find_col('DID',row_max,col_max_num)
print(DID_col)
print(DID_row_1st)


print('\n')
end_time = time.time()
print(str(end_time-start_time)+' s')