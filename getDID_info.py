'''
16/08/203
Pham Minh Nhat
PNM3HC
Get DID info from excel sheet

'''



import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import os
import re


D_ID = 'DID'
ramcell = 'measurement'


start_time = time.time()
cwd = os.getcwd()

def find_ws(keyname):
    sheet_name = ''
    for i in wbl.sheetnames :
        print(i)
        if(keyname in i):
            sheet_name = i
    return sheet_name


def Find_file_Name(dir,keyname):
    fileNames = os.listdir(dir)
    for fileName in fileNames:
        if(keyname in fileName):
            file_name = fileName

    return file_name

def Find_cell(keyname,row_max,col_max_num):
    col_return = 0
    row_return = 0
    row_s = row_max
    col_s = col_max_num
    print(row_s,col_s)
    for row in range(1,row_max+1):
        for col in range(1,col_max_num+1):
            col_char = get_column_letter(col)
            if keyname in str(wsl[col_char +str(row)].value):
                col_return = col
                row_return = row
    print('[' + str(get_column_letter(col_return))+str(row_return) + '] : '+wsl[get_column_letter(col_return)+str(row_return)].value)
    return col_return,row_return

def get_cell(extract_cell):
    for i in range (len(extract_cell)):
        if extract_cell[i] == '.':
            cell_c = extract_cell[i+1:-1]
            break
    num = re.findall(r'\d+', cell_c)
    char = cell_c.replace(num[0],'')
    row_max = int(num[0])
    col_max = int(column_index_from_string(char))
    return row_max, col_max
#Access Excel Workbook

File_dir = Find_file_Name(cwd,'DID_sheet')
print(File_dir)
print('Accessing '+str(File_dir))
wbl = load_workbook(File_dir)
ws_name = find_ws('Cal')
if ws_name != '':
    wsl = wbl[ws_name]
else:
    wsl = wbl.active
print('accessing: '+ str(wsl))
# print(wsl['E'][104].value)
extract_cell = str(wsl[1][-1])
row_max, col_max = get_cell(extract_cell)
extract_cell = str(wsl[get_column_letter(col_max)][-1])
row_max, col_max = get_cell(extract_cell)
print(row_max, col_max)


DID_col,DID_row_1st = Find_cell(D_ID,row_max,col_max)
print(DID_col,DID_row_1st)

print('\n')
end_time = time.time()
print(str(end_time-start_time)+' s')