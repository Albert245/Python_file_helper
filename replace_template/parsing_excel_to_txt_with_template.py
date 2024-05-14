import openpyxl
import re
from openpyxl.utils import get_column_letter, column_index_from_string

def get_cell(extract_cell):
    for i in range (len(extract_cell)):
        if extract_cell[i] == '.':
            cell_c = extract_cell[i+1:-1]
            break
    num = re.findall(r'\d+', cell_c)
    char = cell_c.replace(num[0],'')
    row_max = int(num[0])
    col_max = int(column_index_from_string(char))
    return row_max, col_max


output_text = ''

template_file_name = str(input('template file (txt file): '))
excel_file_name = str(input('Excel file name : '))
start_row = int(input('Start row : '))


lines = []
specific_string = "stop"

# Read the template file
with open(template_file_name, 'r') as template_file:
    template = str(template_file.read())




# Read the Excel file and extract data
workbook = openpyxl.load_workbook(excel_file_name)
sheet = workbook.active

extract_cell = str(sheet[1][-1])
row_max, col_max = get_cell(extract_cell)
extract_cell = str(sheet[get_column_letter(col_max)][-1])
row_max, col_max = get_cell(extract_cell)





while True:
    line = input("Enter a mapping '[substring] [Excel column]' Ex: 'DID A' (type 'stop' to end): ")
    if line == specific_string:
        break
    line_values = line.split()
    lines.append(line_values)

for row in range (start_row,row_max):
    temp = "\n" + template
    for i in range(len(line)):
        col = line[i][1]
        # logic here
        temp_val = str(sheet[col][row].value)
        line[i].append(temp_val)
        temp = temp.replace(line[i][0], line[i][2])

    output_text = output_text + temp


    # DID = str(sheet['A'][row].value)
    # AVSID = str(sheet['L'][row].value)
    # POS = str(sheet['C'][row].value)
    # content = str(sheet['B'][row].value)
    # if POS != '1':
    #     POS = '1...' + POS
    # HEXRANGE = '0x' + str(sheet['F'][row].value) + ' ~ 0x' +str(sheet['G'][row].value)
    # PHYSRANGE = str(sheet['H'][row].value) + ' ~ ' + str(sheet['I'][row].value)
    # UNIT = str(sheet['J'][row].value)

    # temp = template.replace('{{DID}}', str(DID))
    # temp = temp.replace('AVSID', str(AVSID))
    # temp = temp.replace('@POS', str(POS))
    # temp = temp.replace('@UNIT', str(UNIT))
    # temp = temp.replace('HEXRANGE', str(HEXRANGE))
    # temp = temp.replace('PHYSRANGE', str(PHYSRANGE))
    # temp = temp.replace('{{content}}', str(content))
    # output_text = output_text + "\n" + temp



# Write the modified content to a new text file
with open('output.txt', 'w') as output_file:
    output_file.write(output_text)

print("File created successfully!")