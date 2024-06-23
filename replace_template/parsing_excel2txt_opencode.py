import openpyxl
import re
from openpyxl.utils import get_column_letter, column_index_from_string

def get_cell(extract_cell):
    for i in range (len(extract_cell)):
        if extract_cell[i] == '.':
            cell_c = extract_cell[i+1:-1]
            break
    num = re.findall(r'\d+', cell_c)
    char = cell_c.replace(num[0],'')
    row_max = int(num[0])
    col_max = int(column_index_from_string(char))
    return row_max, col_max

# Read the template file

with open('template.txt', 'r') as template_file:
    template = str(template_file.read())

# Read the Excel file and extract data
workbook = openpyxl.load_workbook('data.xlsx')
sheet = workbook.active

extract_cell = str(sheet[1][-1])
row_max, col_max = get_cell(extract_cell)
extract_cell = str(sheet[get_column_letter(col_max)][-1])
row_max, col_max = get_cell(extract_cell)

output_text = ''

for row in range (1,row_max):


    DID = str(sheet['A'][row].value)
    AVSID = str(sheet['L'][row].value)
    POS = str(sheet['C'][row].value)
    content = str(sheet['B'][row].value)
    if POS != '1':
        POS = '1...' + POS
    PHYSRANGE = str(sheet['H'][row].value) + ' ~ ' + str(sheet['I'][row].value)
    UNIT = str(sheet['J'][row].value)
    FACTOR = str(sheet['E'][row].value)

    temp = template.replace('{{DID}}', str(DID))
    temp = temp.replace('{{AVSID}}', str(AVSID))
    temp = temp.replace('{{POS}}', str(POS))
    temp = temp.replace('{{UNIT}}', str(UNIT))
    temp = temp.replace('{{FACTOR}}', str(FACTOR))
    temp = temp.replace('{{PHYS}}', str(PHYSRANGE))
    temp = temp.replace('{{CONTENT}}', str(content))
    if row == 1:
        output_text = temp
    else:
        output_text = output_text + "\n" + temp



# Write the modified content to a new text file
with open('output.txt', 'w') as output_file:
    output_file.write(output_text)

print("\nFile created successfully!")